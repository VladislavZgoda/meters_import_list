import re
from pathlib import Path
from typing import TypedDict

from openpyxl import load_workbook, Workbook
from openpyxl.utils import column_index_from_string

from process_files import ImportData

dynamic_columns_ty = {
    "consumer_code": "AX",
    "serial_number": "AE",
    "address": "AB",
    "tp_number": "T",
}

static_columns_ty = {
    "AG": "01.01.2026",
    "AY": "Тимашевск",
    "AZ": "2 тарифа",
    "AJ": 3,
    "X": "УСПД",
    "Y": "яч. ф-н/д 0,4 кВ",
    "Z": "Ячейка присоединения",
    "AA": "ф-н/д",
    "B": 'АО "Электросети Кубани"',
    "C": "Тимашевскэлектросеть",
}


def fill_in_import_list(import_data: list[ImportData]) -> Workbook:
    script_dir = Path(__file__).resolve().parent
    import_list_path = script_dir / "template" / "import_list.xlsx"

    wb = load_workbook(import_list_path)
    ws_ty = wb["ТУ"]
    ws_fl = wb["ФЛ"]
    ws_ul = wb["ЮЛ"]

    ws_row = 4

    for i, record in enumerate(import_data):
        for key, col_letter in dynamic_columns_ty.items():
            col_num = column_index_from_string(col_letter)
            ws_ty.cell(row=ws_row, column=col_num, value=record.get(key))

        for col_letter, value in static_columns_ty.items():
            col_num = column_index_from_string(col_letter)
            ws_ty.cell(row=ws_row, column=col_num, value=value)

        ws_ty.cell(
            row=ws_row,
            column=column_index_from_string("A"),
            value=i + 1,
        )

        ws_ty.cell(
            row=ws_row,
            column=column_index_from_string("AD"),
            value=select_device_type(record["device_type"]),
        )

        ws_row += 1

    ws_row = 3
    row_num = 1

    for record in import_data:
        if record["consumer_code"].startswith("230700"):
            ws_fl.cell(
                row=ws_row,
                column=column_index_from_string("A"),
                value=row_num,
            )
            ws_fl.cell(
                row=ws_row,
                column=column_index_from_string("G"),
                value=record["consumer_code"],
            )

            parsed_name = parse_consumer_name(record["consumer_name"])

            if parsed_name is None:
                ws_fl.cell(
                    row=ws_row,
                    column=column_index_from_string("B"),
                    value=record["consumer_name"],
                )
            else:
                ws_fl.cell(
                    row=ws_row,
                    column=column_index_from_string("B"),
                    value=parsed_name["surname"],
                )
                ws_fl.cell(
                    row=ws_row,
                    column=column_index_from_string("C"),
                    value=parsed_name["name"],
                )
                ws_fl.cell(
                    row=ws_row,
                    column=column_index_from_string("D"),
                    value=parsed_name["patronymic"],
                )
            ws_row += 1
            row_num += 1

    ws_row = 3
    row_num = 1

    for record in import_data:
        if not record["consumer_code"].startswith("230700"):
            ws_ul.cell(
                row=ws_row,
                column=column_index_from_string("A"),
                value=row_num,
            )
            ws_ul.cell(
                row=ws_row,
                column=column_index_from_string("B"),
                value=row_num,
            )
            ws_ul.cell(
                row=ws_row,
                column=column_index_from_string("F"),
                value=record["consumer_code"],
            )
            ws_row += 1
            row_num += 1
    return wb


def select_device_type(type: str) -> str:
    if type.startswith("AD11S"):
        return "Матрица - AD11S"
    elif type.startswith("AD13S"):
        return "Матрица - AD13S"
    elif type.startswith("AD11A"):
        return "Матрица - AD11A"
    elif type.startswith("AD13A"):
        return "Матрица - AD13A"
    else:
        return type


class SurnameWithInitials(TypedDict):
    surname: str
    name: str
    patronymic: str


def parse_consumer_name(name: str) -> SurnameWithInitials | None:
    # Фамилия (не пробельные символы), пробел, инициал (одна буква + возможно точка),
    # пробел, инициал (одна буква + возможно точка)
    match = re.match(r"^(\S+)\s+(\S)\.?\s+(\S)\.?$", name.strip())

    if not match:
        return None

    return {
        "surname": match.group(1),
        "name": match.group(2) + ".",
        "patronymic": match.group(3) + ".",
    }
